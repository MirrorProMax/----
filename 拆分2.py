import numpy as np
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import SubFunc
import random


def refreshConsole():
    """清理控制台"""
    os.system("clear")
    print(f"{datetime.now()}\n")


def loadExcelFile(filename):
    """加载 Excel 文件"""
    try:
        workbook = load_workbook(filename=filename)
        return workbook.active if workbook else None
    except Exception as e:
        print(f"Error: 无法加载文件 {filename}. 错误信息: {e}")
        return None


def randomSplitTotal(totalAmount: float):
    # 参数校验 🔧
    if totalAmount is None:
        return None
    if not isinstance(totalAmount, (float, int)):
        raise TypeError("总金额必须是数值类型")
    if totalAmount <= 0:
        raise ValueError("总金额必须大于0")

    # 随机生成3-6个分组
    groupCount: int = int(np.random.randint(3, 7))

    # 生成金额分配权重（确保总和精确）
    fractionsNdArray = np.random.dirichlet(np.ones(groupCount), size=1)
    splitAmountsNdArray = (fractionsNdArray * totalAmount).flatten()

    splitAmounts = splitAmountsNdArray.tolist()

    result: list = []
    remaining: float = float(totalAmount)  # 剩余待分配金额

    unitPriceLowerLimit = 26
    unitPriceUpperLimit = 28

    for i in range(groupCount):
        # 动态调整最后一组逻辑
        if i == groupCount - 1:
            currentAmount: float = float(remaining)
        else:
            currentAmount: float = splitAmounts[i]

        # 生成单价（严格限制范围）
        currentUnitPrice = round(
            SubFunc.randomFloat(unitPriceLowerLimit, unitPriceUpperLimit), 2
        )

        # 计算数量（保留两位小数）
        quantity: float = float(round(currentAmount / currentUnitPrice, 2))

        # 记录结果
        result.append(
            [
                i + 1,
                currentUnitPrice,
                quantity,
                currentAmount,
                remaining - currentAmount,  # 新增监控字段
            ]
        )

        # 更新剩余金额（最后一组自动归零）
        remaining -= float(currentAmount)

    # 强制校验总和
    if abs(sum(splitAmounts) - totalAmount) > 0.01:
        raise ValueError("金额分配不准确")

    return result


def randomDay(originalDate: datetime) -> datetime:
    # 随机生成“日”的值（1到28之间，避免月份天数问题）
    if isinstance(originalDate, datetime):
        randomDay = random.randint(1, 28)
        modifiedDate = originalDate.replace(day=randomDay)
    else:
        modifiedDate = originalDate  # 如果不是日期类型，保持原样
    return modifiedDate


def processRows(inSheet, outputData: list):
    rows = list(inSheet.iter_rows(min_row=1, values_only=True))
    columns = list(inSheet.iter_cols(min_row=1, values_only=True))
    if not columns or not rows:
        print("Error: 表格数据为空或格式不正确。")
        return
    dateRow = rows[0]
    headers = columns[0]

    for rowIndex, row in enumerate(rows[1:], start=1):
        for colIndex, cellValue in enumerate(row[1:], start=1):
            if cellValue is None:
                continue
            splitResult = randomSplitTotal(cellValue)

            # 展平结果并添加主体和日期信息
            for item in splitResult:

                outputData.append(
                    [
                        row[0],  # 主体
                        randomDay(dateRow[colIndex]),  # 日期
                        *item,  # 展平的分组数据
                    ]
                )


def saveToExcel(outputData, filenamePrefix="Output"):
    """保存数据到 Excel 文件"""
    if not outputData:
        print("Warning: 没有数据需要保存。")
        return

    try:
        outWorkbook = Workbook()
        outSheet = outWorkbook.active
        if outSheet:
            outSheet.freeze_panes = "B2"  # 冻结第一行
            outSheet.append(
                ["主体", "日期", "序号", "单价", "数量", "金额", "剩余金额"]
            )
            for row in outputData:
                outSheet.append(row)
            filename = (
                f"{filenamePrefix}_{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
            )
            outWorkbook.save(filename)
            print(f"数据已保存到文件: {filename}")
    except Exception as e:
        print(f"Error: 无法保存文件。错误信息: {e}")


def main():
    refreshConsole()
    sheet = loadExcelFile("育肥牛.xlsx")
    if sheet is None:
        print("Error: 无法加载工作表。")
        return

    outputData = []
    processRows(sheet, outputData)
    saveToExcel(outputData)


if __name__ == "__main__":
    main()
