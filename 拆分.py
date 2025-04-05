import random
import os
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import SubFunc


def clear_console():
    """清理控制台"""
    os.system("clear")
    print(f"{datetime.now()}\n")


def load_excel_file(filename):
    """加载 Excel 文件"""
    try:
        workbook = load_workbook(filename=filename)
        return workbook.active if workbook else None
    except Exception as e:
        print(f"Error: 无法加载文件 {filename}. 错误信息: {e}")
        return None


def generate_random_split(
    remaining_amount, base_amount, volatility, rounding_digits, price_range
):
    """生成随机拆分数据"""
    if remaining_amount <= 0:
        return 0, 0, 0  # 防止负值或无效数据

    random_ratio = SubFunc.randomFloat(下限=-volatility, 上限=volatility)
    if remaining_amount <= base_amount:
        split_amount = round(remaining_amount, rounding_digits)
    else:
        split_amount = round(base_amount * (1 + random_ratio), rounding_digits)

    price = round(SubFunc.randomFloat(*price_range), 1)
    quantity = round(split_amount / price, rounding_digits)
    return split_amount, price, quantity


def process_rows(sheet, output_data):
    """处理 Excel 表格的行数据"""
    try:
        columns = list(sheet.iter_cols(min_row=2, values_only=True))
        rows = list(sheet.iter_rows(min_row=1, values_only=True))
        if not columns or not rows:
            print("Error: 表格数据为空或格式不正确。")
            return

        headers = columns[0]
        date_row = rows[0]

        for row_index, row_data in enumerate(rows[1:], start=1):
            rounding_digits = 2
            for col_index, cell_value in enumerate(row_data[1:], start=1):
                if cell_value is None:
                    continue

                remaining_amount = int(cell_value)
                volatility = 0.1
                price_range = (26, 28)
                target_count_range = (3, 6)
                target_count = random.randint(*target_count_range)
                base_amount = int(remaining_amount / target_count)

                sequence = 1
                while remaining_amount > 0:
                    split_amount, price, quantity = generate_random_split(
                        remaining_amount,
                        base_amount,
                        volatility,
                        rounding_digits,
                        price_range,
                    )
                    if split_amount <= 0:  # 防止死循环
                        break

                    remaining_amount = round(
                        remaining_amount - split_amount, rounding_digits
                    )

                    # 检查日期有效性
                    date_value = date_row[col_index]
                    if not isinstance(date_value, datetime):
                        date_value = datetime.now()  # 默认使用当前日期

                    output_row = [
                        headers[row_index - 1],
                        date_value + timedelta(days=SubFunc.randomInt(0, 20)),
                        sequence,
                        price,
                        quantity,
                        split_amount,
                    ]
                    output_data.append(output_row)
                    sequence += 1
    except Exception as e:
        print(f"Error: 处理行数据时发生错误。错误信息: {e}")


def save_to_excel(output_data, filename_prefix="Output"):
    """保存数据到 Excel 文件"""
    if not output_data:
        print("Warning: 没有数据需要保存。")
        return

    try:
        workbook = Workbook()
        sheet = workbook.active
        if sheet:
            sheet.append(["抬头", "日期", "序号", "单价", "数量", "金额"])
            for row in output_data:
                sheet.append(row)
            filename = (
                f"{filename_prefix}_{datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.xlsx"
            )
            workbook.save(filename)
            print(f"数据已保存到文件: {filename}")
    except Exception as e:
        print(f"Error: 无法保存文件。错误信息: {e}")


def main():
    clear_console()
    sheet = load_excel_file("育肥牛.xlsx")
    if sheet is None:
        print("Error: 无法加载工作表。")
        return

    output_data = []
    process_rows(sheet, output_data)
    save_to_excel(output_data)


if __name__ == "__main__":
    main()
